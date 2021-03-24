#import ipcalc
import sys
import os
import openpyxl
import ipaddress

def removeSpaces(string): 
    return "".join(string.split())


files=[]

for file in os.listdir('ExcelSources'):
    if '.xlsx' in file:
        files.append('ExcelSources'+'\\'+file)

for f in files:
    print(f)
    wb=openpyxl.load_workbook(f)
    #worksheet=wb.get_sheet_by_name('Structure.xls')
    worksheet=wb.get_sheet_by_name(wb.get_sheet_names()[0])
    ipSubnetColumn= worksheet['E']
    
    for ipSubnet in range ( 3,len(ipSubnetColumn)+1) :
        rangeIP=worksheet.cell(row=ipSubnet, column=5).value
        print(rangeIP)
        #subnet=ipcalc.Network(rangeIP.strip())
        
        try:
            subnet=ipaddress.ip_network(removeSpaces(rangeIP))
            worksheet.cell(ipSubnet,10).value=str(subnet[0])+"-"+str(subnet[-1])
            worksheet.cell(ipSubnet,11).value=str(subnet.num_addresses)
        except ValueError:
            worksheet.cell(ipSubnet,10).value="Invalid IP"
            worksheet.cell(ipSubnet,11).value="find the correct ip"


        
        
    print("Finished processing file "+f)
    wb.save(f)
    


