import ipaddress
import openpyxl
poop=ipaddress.ip_network("10.10.0.0/24")
print(poop.num_addresses)


wb=openpyxl.load_workbook("CoreNetwork.xlsx")
print(wb.sheetnames)
        
#worksheet=wb.get_sheet_by_name(wb.sheetnames()[0])