import openpyxl

class excelworld:

    def __init__(self,fileLocation):
        self.fileLocation=fileLocation
        self.wb = openpyxl.load_workbook(self.fileLocation)
        self.first_sheet = self.wb.get_sheet_names()[0]
        self.worksheet = self.wb.get_sheet_by_name(self.first_sheet)


    def getAllHostNamesFromExcel(self):
        first_column = self.worksheet['B']
        #print(firs)
        #list2 = [x for x in first_column if x.value != None]
        for x in first_column:
            if(x.value != None):
                x.value=x.value.split(".", 1)[0]

        return first_column

    def saveAllBlimpDetails(self,blimpArray):
        jeher=2
        self.worksheet.cell(row=1,column=3).value="Application Name"
        self.worksheet.cell(row=1,column=4).value="Responsible Engg"
        self.worksheet.cell(row=1,column=5).value="Responsible Ops"
        for blimps in blimpArray:
            print(blimps[1])
            self.worksheet.cell(row=jeher,column=3).value=blimps[0]
            self.worksheet.cell(row=jeher,column=4).value=blimps[1]
            self.worksheet.cell(row=jeher,column=5).value=blimps[2]
            jeher += 1
        self.wb.save(self.fileLocation)