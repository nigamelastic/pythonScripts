folderPath = r'path to folder'
nameOfExcel = r'folder name'

import sys
import os
import openpyxl

wb_obj = openpyxl.Workbook()
sheet_obj = wb_obj.active
sheet_obj.cell(row = 1, column = 1).value = 'Item Nr.'
sheet_obj.cell(row = 1, column = 2).value = 'link'

i = 2
for file in os.listdir(folderPath):
   if '.pdf' in file:
       sheet_obj.cell(row = i, column = 1).value = file[:8]
       sheet_obj.cell(row = i, column = 2).value = folderPath+'\\'+file
       i+=1


wb_obj.save(nameOfExcel)    