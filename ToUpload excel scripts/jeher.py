import sys
import os
import xlrd
from selenium.common.exceptions import *
import configparser
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium import webdriver
import time
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.common.action_chains import ActionChains
import openpyxl
files = []
excelfoldername = r'pathexcel'
for file in os.listdir(excelfoldername):
   if '.xlsx' in file:
       files.append(excelfoldername+'\\'+file)


src =r'path.xlsx'
binary = r'pathfirefox.exe'
options = Options()
options.set_headless(headless=False)
options.binary = binary

cap = DesiredCapabilities().FIREFOX
cap["marionette"] = True 
fp = webdriver.FirefoxProfile(r'default')

driver = webdriver.Firefox(fp, firefox_options=options,  capabilities=cap, executable_path="geckodriver.exe")

driver.get("https://vendorcentral.amazon.de/hz/vendor/members/shipment-mgr/pendingasnselect")



driver.implicitly_wait(20)
for f in files:
    wb_obj = openpyxl.load_workbook(f) 
    sheet_obj = wb_obj.active
    workbook = xlrd.open_workbook(f)
    sheet= workbook.sheet_by_index(0)
    PO = sheet.cell(0,3).value
    kartonSTKValues= sheet.col_values(5)[2:]
    artikelNum= sheet.col_values(1)[2:]
    pallets = sheet.cell(2,7).value
    menge = sheet.col_values(3)[2:]
   
    mHDs = sheet.col_values(8)[2:]
    sendungsnummer= sheet.cell(0,1).value

    driver.get("https://vendorcentral.amazon.de/hz/vendor/members/shipment-mgr/pendingasnselect")
    time.sleep(5)
    driver.find_element_by_css_selector('#tab-controller > kat-tab-pane:nth-child(1) > kat-tab-header:nth-child(2)').click()
    time.sleep(5)
    driver.find_element_by_xpath('//*[@id="table-search-input"]/input').send_keys(PO)
    
    chkbox=WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="prepaid-po-select-table"]//kat-link[contains(@label,"'+PO+'")]/../preceding-sibling::td/kat-checkbox/div')))
    handlercode=driver.find_element_by_xpath('//*[@id="prepaid-po-select-table"]//kat-link[contains(@label,"'+PO+'")]/../../td[6]').text
    bestelldatum=driver.find_element_by_xpath('//*[@id="prepaid-po-select-table"]//kat-link[contains(@label,"'+PO+'")]/../../td[3]').text
    versandbis=driver.find_element_by_xpath('//*[@id="prepaid-po-select-table"]//kat-link[contains(@label,"'+PO+'")]/../../td[4]').text
    
    
    chkbox.click()
    driver.find_element_by_xpath('//*[@id="next-button"]/button').click()
    wait = WebDriverWait(driver, 10, poll_frequency=1, ignored_exceptions=[ElementNotVisibleException, ElementNotSelectableException])
       
    if(handlercode == ""):
        radiobtn = driver.find_element_by_xpath("//*[@id='katal-id-2']")
        #radiobtn = wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='katal-id-2']")))
        driver.execute_script('arguments[0].click()',radiobtn)
        #radiobtn.click()
        driver.find_element_by_xpath('//*[@id="katal-id-0"]').send_keys(sum(map(int,kartonSTKValues)))
        driver.find_element_by_xpath('//*[@id="noAmznccRadioGroup"]/div[2]/kat-radiobutton[2]/div/div[1]/span').click()

        driver.find_element_by_xpath('//*[@id="hmd2f-exit"]').click()

        driver.find_element_by_xpath('//*[@id="next-button"]/button').click()
        driver.find_element_by_xpath('//*[@id="katal-id-12"]').send_keys(bestelldatum)
        driver.find_element_by_xpath('//*[@id="katal-id-13"]').send_keys(versandbis)
        #driver.find_element_by_xpath('//*[@id="hmd2f-exit"]').click()
        driver.find_element_by_xpath('//*[@id="katal-id-1"]').send_keys('0')
        driver.find_element_by_xpath('//*[@id="katal-id-2"]').send_keys(pallets)
        
        driver.find_element_by_xpath('/html/body/div[2]/div[2]/div/form/div/div[9]/section/kat-dropdown/div/div[1]').click()
        time.sleep(2)
       
        selectdropdown = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@data-value="DEUTP"]')))
        selectdropdown.click()
        driver.find_element_by_xpath('//*[@id="katal-id-6"]').send_keys(sendungsnummer)
    
        
        
        #wait.until(EC.invisibility_of_element_located((By.XPATH,"//a[@class='a-link-normal']")))
        driver.find_element_by_xpath('//*[@id="submit-details-button"]/button').click() 
    
        for aNo in artikelNum:
            driver.find_element_by_xpath('//*[@id="asn-shipment-items-table"]//td[contains(text(),"'+aNo+'")]/..//div[@class="checkbox"]').click()

        driver.find_element_by_xpath('//*[@id="next-button"]').click()
        driver.find_element_by_xpath('//*[@id="submit-button"]').click()
    else:
        driver.find_element_by_xpath('//*[@id="katal-id-0"]').send_keys(sum(map(int,kartonSTKValues)))
        driver.find_element_by_xpath('//*[@id="next-button"]/button').click()
        time.sleep(2)
    
        
        totalStk=0
        i=0

        while i < len(artikelNum):
            initialValue=totalStk+1
            print('________________________________________'+artikelNum[i])
            driver.find_element_by_xpath('(//*[@id="shipment-items-sscc-table"]//td[contains(text(),"'+artikelNum[i]+'")]/../td/kat-dropdown/div)[1]').click()
            containerCodes = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="shipment-items-sscc-table"]//td[contains(text(),"'+artikelNum[i]+'")]/../td/kat-dropdown/div/div/div/div[contains(@data-value,"multiple_cartons")]/div')))
            containerCodes.click()
            
            bearbiten = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="shipment-items-sscc-table"]//td[contains(text(),"'+artikelNum[i]+'")]/../td[10]/div/span/kat-link/a')))
            bearbiten.click()
            
            
            for j in range(int(kartonSTKValues[i])):
                
                totalStk=totalStk+1
                
                caspackUnit=int(int(menge[i])/int(kartonSTKValues[i]))
                driver.find_element_by_xpath('(//*[@id="multiple-carton-data-table"]/table/tbody/tr/td[text()="'+str(totalStk)+'"]/following-sibling::td/kat-input/input)[1]').send_keys(str(caspackUnit))
                if mHDs[i]:
                    driver.find_element_by_xpath('//*[@id="multiple-carton-data-table"]/table/tbody/tr/td[text()="'+str(totalStk)+'"]/following-sibling::td/kat-date-picker/div/kat-input-group/kat-input/input').send_keys(mHDs[i])
                
                #updatedFile.get_sheet(0).write(i+2,6,str(totalStk+1)+'-'+str(totalStk+int(kartonSTKValues[i])))
            updatedCell=sheet_obj.cell(row = i+3, column =7)
            updatedCell.value=str(initialValue)+'-'+str(totalStk)
            
            driver.find_element_by_xpath('//*[@id="multiple-carton-popup-save"]/button').click()
            i += 1
        
        driver.find_element_by_xpath('//*[@id="next-button"]').click()

        driver.find_element_by_xpath('//*[@id="katal-id-12"]').send_keys(bestelldatum)
        driver.find_element_by_xpath('//*[@id="katal-id-13"]').send_keys(versandbis)
        #driver.find_element_by_xpath('//*[@id="hmd2f-exit"]').click() 
        driver.find_element_by_xpath('//*[@id="katal-id-1"]').send_keys('0')
        driver.find_element_by_xpath('//*[@id="katal-id-2"]').send_keys(pallets)
        driver.find_element_by_xpath('/html/body/div[2]/div[2]/div/form/div/div[9]/section/kat-dropdown/div/div[1]').click()
        time.sleep(2)
       
        selectdropdown = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@data-value="DEUTP"]')))
        selectdropdown.click()
        driver.find_element_by_xpath('//*[@id="katal-id-6"]').send_keys(sendungsnummer)
        #driver.find_element_by_xpath('//*[@id="hmd2f-exit"]').click() 
        driver.find_element_by_xpath('//*[@id="submit-details-button"]/button').click() 
       
        driver.find_element_by_xpath('//*[@id="submit-button"]').click()
        #driver.find_element_by_xpath('/html/body/div[2]/div[2]/div/div[5]/section/kat-button/button').click()
        
    wb_obj1 = openpyxl.load_workbook(src) 
    sheet_obj1 = wb_obj1.active
    sheet_obj1.cell(row=7, column=1).value = driver.find_element_by_xpath('/html/body/div[2]/div[2]/div/div[2]/div[2]/div/section[2]/div[2]/section[2]').text
    sheet_obj1.cell(row=7, column=7).value = driver.find_element_by_xpath('/html/body/div[2]/div[2]/div/div[2]/div[2]/div/section[1]/div[1]/section[2]').text
    sheet_obj1.cell(row=8, column=7).value = sendungsnummer
    sheet_obj1.cell(row=10, column=7).value = PO
    wb_obj1.save("template\\"+sendungsnummer+".xlsx")
    wb_obj.save(f)
    if(handlercode != ""): 
        link1=WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="displayShipmentButtons"]/div[1]/section[2]/span/kat-link/a')))
        link1.click()
        link2=WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="print-label-modal-span"]/kat-button/button')))
        link2.click()
        time.sleep(3)
        driver.switch_to.window(driver.window_handles[1])
        time.sleep(60)
        driver.close()
        driver.switch_to.window(driver.window_handles[0])
driver.quit()