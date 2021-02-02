import configparser
import os
from selenium.webdriver.common.action_chains import ActionChains
import sys
from selenium.common.exceptions import UnexpectedAlertPresentException
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import *
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium import webdriver
import time
from selenium.webdriver.support.ui import Select
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.common.action_chains import ActionChains
import openpyxl
from selenium.common.exceptions import NoSuchElementException   


    
config = configparser.ConfigParser()
foldername = r'pathDHLExcel'
binary = r'pathfirefox.exe'
options = Options()
options.set_headless(headless=False)
options.binary = binary
config.read("happy.config")
username = config.get('DHL','username')
password = config.get('DHL','password')
cap = DesiredCapabilities().FIREFOX
cap["marionette"] = True 
fp = webdriver.FirefoxProfile(r'.default')

driver = webdriver.Firefox(fp, firefox_options=options,  capabilities=cap, executable_path="geckodriver.exe")
driver.get("https://www.ax4.com/ax4/control/man_dispatch_overview_shipper")


def check_exists_by_xpath(xpath):
    try:
        driver.find_element_by_xpath(xpath)
    except NoSuchElementException:
        return False
    return True


driver.implicitly_wait(20)
driver.find_element_by_xpath('//*[@id="loginPage18"]/div/div[2]/form/input').send_keys(username)
driver.find_element_by_xpath('//*[@id="loginPage18"]/div/div[2]/form/div[2]/input').send_keys(password)
driver.find_element_by_xpath('//*[@id="loginPage18"]/div/div[2]/form/button').click()
files = []
excelfoldername = r'pathDHLExcel'
for file in os.listdir(excelfoldername):
   if '.xlsx' in file:
       files.append(excelfoldername+'\\'+file)
for f in files:
    print(f)
    wb_obj = openpyxl.load_workbook(f) 
    sheet_obj = wb_obj.active

    driver.get('https://www.ax4.com/ax4/control/openForm_A_1')
    driver.find_element_by_xpath('//*[@id="ConsigneeAutoComplete"]').send_keys(sheet_obj.cell(row = 13, column =1).value)
    time.sleep(2)
    driver.find_element_by_xpath('//*[@id="ConsigneeAutoComplete"]').send_keys(Keys.RETURN)
    time.sleep(2)
    driver.find_element_by_xpath('//*[@id="CustomersReference"]').send_keys('ASN '+str(sheet_obj.cell(row = 7, column =7).value))
    possss=str(sheet_obj.cell(row = 10, column =7).value)
    print('_________________________________')
    driver.find_element_by_xpath('//*[@id="ConsignorReference"]').send_keys('PO '+ possss.replace('-','- PO'))

    driver.find_element_by_xpath('//*[@id="DeliveryInstruction"]').send_keys('Liefertermin '+str(sheet_obj.cell(row = 22, column =7).value))
    
    driver.find_element_by_xpath('//*[@id="Next"]').click()
    warenbezeichnung=WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="GoodsDescription"]')))
    warenbezeichnung.send_keys('Autozubehör')
    driver.find_element_by_xpath('//*[@id="MarksNumbers"]').send_keys(str(sheet_obj.cell(row = 8, column =7).value))
    counter=0
    karton=sheet_obj.cell(row = 36, column =6).value
    EWP=sheet_obj.cell(row = 36, column =7).value
    EUP=sheet_obj.cell(row = 36, column =8).value
    weight=0
    weight2=0
    weight3=0
    
    if(karton != None):
        
        firstEntry=WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="Positionsó'+str(counter)+'óQuantity"]')))
        firstEntry.send_keys(str(karton))
        weight = sheet_obj.cell(row = 39, column = 6).value
        lange = sheet_obj.cell(row = 36, column = 2).value
        britte = sheet_obj.cell(row = 36, column = 3).value
        hohe = sheet_obj.cell(row = 36, column = 4).value
        
        driver.find_element_by_xpath('//*[@id="Positionsó'+str(counter)+'óPackaging-button"]/span[1]').click()
        dropdownSelect11=WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.XPATH, '(//li[text()="Karton"])['+str(counter+1)+']')))
        time.sleep(2)
        last_height = driver.execute_script("return document.body.scrollHeight")
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        hover = ActionChains(driver).move_to_element(dropdownSelect11)
        hover.perform()
        dropdownSelect11.click()



        driver.find_element_by_xpath('//*[@id="Positionsó'+str(counter)+'óGrossWeight"]').send_keys(str(weight))
        driver.find_element_by_xpath('//*[@id="Positionsó'+str(counter)+'óLength"]').send_keys(str(lange))
        driver.find_element_by_xpath('//*[@id="Positionsó'+str(counter)+'óWidth"]').send_keys(str(britte))
        driver.find_element_by_xpath('//*[@id="Positionsó'+str(counter)+'óHeight"]').send_keys(str(hohe))
        counter=counter+1

    if(EWP != None):
        if(counter > 0):
            driver.find_element_by_xpath('//*[@id="AddPosition"]').click()

        SecondEntry=WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="Positionsó'+str(counter)+'óQuantity"]')))
        SecondEntry.send_keys(str(EWP))
        weight2 = sheet_obj.cell(row = 39, column = 7).value
        lange2 = sheet_obj.cell(row = 37, column = 2).value
        britte2 = sheet_obj.cell(row = 37, column = 3).value
        hohe2 = sheet_obj.cell(row = 37, column = 4).value



        driver.find_element_by_xpath('//*[@id="Positionsó'+str(counter)+'óPackaging-button"]/span[1]').click()
        dropdownSelect2=WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.XPATH, '(//li[text()="Einwegpalette"])['+str(counter+1)+']')))
        time.sleep(2)
        last_height = driver.execute_script("return document.body.scrollHeight")
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        hover = ActionChains(driver).move_to_element(dropdownSelect2)
        hover.perform()
        dropdownSelect2.click()

        
        driver.find_element_by_xpath('//*[@id="Positionsó'+str(counter)+'óGrossWeight"]').send_keys(str(weight2))
        driver.find_element_by_xpath('//*[@id="Positionsó'+str(counter)+'óLength"]').send_keys(str(lange2))
        driver.find_element_by_xpath('//*[@id="Positionsó'+str(counter)+'óWidth"]').send_keys(str(britte2))
        driver.find_element_by_xpath('//*[@id="Positionsó'+str(counter)+'óHeight"]').send_keys(str(hohe2))
        counter=counter+1

    if(EUP != None):
        if(counter > 0):
            driver.find_element_by_xpath('//*[@id="AddPosition"]').click()
        thirdEntry=WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="Positionsó'+str(counter)+'óQuantity"]')))
        thirdEntry.send_keys(str(EUP))
        weight3 = sheet_obj.cell(row = 39, column = 8).value
        lange3 = sheet_obj.cell(row = 38, column = 2).value
        britte3 = sheet_obj.cell(row = 38, column = 3).value
        hohe3 = sheet_obj.cell(row = 38, column = 4).value
        #driver.find_element_by_xpath('//*[@id="Positionsó'+str(counter)+'óPackaging"]').click()
        driver.find_element_by_xpath('//*[@id="Positionsó'+str(counter)+'óPackaging-button"]/span[1]').click()
        dropdownSelect1=WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.XPATH, '(//li[text()="Europalette"])['+str(counter+1)+']')))
        time.sleep(2)
        last_height = driver.execute_script("return document.body.scrollHeight")
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        hover = ActionChains(driver).move_to_element(dropdownSelect1)
        hover.perform()
        dropdownSelect1.click()
        driver.find_element_by_xpath('//*[@id="Positionsó'+str(counter)+'óGrossWeight"]').send_keys(str(weight3))
        driver.find_element_by_xpath('//*[@id="Positionsó'+str(counter)+'óLength"]').send_keys(str(lange3))
        driver.find_element_by_xpath('//*[@id="Positionsó'+str(counter)+'óWidth"]').send_keys(str(britte3))
        driver.find_element_by_xpath('//*[@id="Positionsó'+str(counter)+'óHeight"]').send_keys(str(hohe3))
        counter=counter+1
    




    driver.find_element_by_xpath('//*[@id="Next"]').click()
    time.sleep(2)
    dropselect1=WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="Service-button"]')))
    hover = ActionChains(driver).move_to_element(dropselect1)
    hover.perform()
    dropselect1.click()
    

    
    sumTotalweight= weight+weight2+weight3
    if(sumTotalweight<2500):
        euroConnect=WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.XPATH, '//li[text()="Euroconnect"]')))
        hover = ActionChains(driver).move_to_element(euroConnect)
        hover.perform()
        euroConnect.click()
    else:
        euroline=WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.XPATH, '//li[text()="Euroline"]')))
        hover = ActionChains(driver).move_to_element(euroline)
        hover.perform()
        euroline.click()
    
    driver.find_element_by_xpath('//*[@id="Next"]').click()
    

    abholdatum=WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ReceiptDate"]')))
    
    

    checkboxXpath="//*[@id='Ekaer']"
    existsValue=check_exists_by_xpath(checkboxXpath)

    if(existsValue):
        driver.find_element_by_xpath(checkboxXpath).click()
    driver.find_element_by_xpath("//*[@id='SpecialInstructions']").send_keys("CARP-System")

    #time.sleep(10)
    driver.find_element_by_xpath('//*[@id="Next"]').click()


    saveButton1=WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="SaveButton"]')))
    saveButton1.click()

    saveButtonConfirmation=WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="openFormShowPageMsg"]/div/div/table/tbody/tr/td[2]/button')))
    saveButtonConfirmation.click()
    time.sleep(3)
driver.quit()    


    